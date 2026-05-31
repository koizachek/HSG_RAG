"""
UAT branch comparison runner.

This file is intentionally usable both as:
- a direct script: python tests/test_uat_branch_comparison.py --excel /path/UAT.xlsx
- an opt-in pytest test: RUN_UAT_BRANCH_COMPARISON=1 pytest tests/test_uat_branch_comparison.py -s

It compares the three agreed chatbot states:
- main: legacy long prompt, multi-agent
- fix/chatbot-overhaul: short RAG-first prompt, single-agent
- test/overhaul-multi-agent: short RAG-first prompt, multi-agent

The runner creates temporary git worktrees so branch imports stay isolated.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import textwrap
import time
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pytest

try:
    from langsmith.run_trees import RunTree as LangSmithRunTree
except Exception:
    LangSmithRunTree = None


DEFAULT_EXCEL = Path(__file__).resolve().parent / "fixtures" / "UAT.xlsx"
DEFAULT_BRANCHES = [
    ("main_legacy_multi", "origin/main"),
    ("overhaul_single", "origin/fix/chatbot-overhaul"),
    ("overhaul_multi", "origin/test/overhaul-multi-agent"),
]

SKIPPED_SHEETS = {"About", "TestProtocol", "Reporting"}
PROGRAMME_COST_TRUTHS = {
    "emba_hsg": {"label": "EMBA HSG", "amount": "77500"},
    "iemba": {"label": "IEMBA", "amount": "85000"},
    "emba_x": {"label": "emba X", "amount": "110000"},
}
STALE_OR_DISCOUNT_COSTS = {
    "72500",  # old EMBA deadline-linked fee
    "80000",  # old IEMBA early registration fee
    "84000",  # old IEMBA final fee in UAT sheet
    "99000",  # old / early emba X fee
}

DEFAULT_JUDGE_MODEL = os.getenv("UAT_JUDGE_MODEL", "gpt-4o-mini")


@dataclass
class UATCase:
    case_id: str
    sheet: str
    title: str
    persona: list[str] = field(default_factory=list)
    user_turns: list[str] = field(default_factory=list)
    expected_bot_behaviour: list[str] = field(default_factory=list)
    success_criteria: list[str] = field(default_factory=list)
    red_flags: list[str] = field(default_factory=list)
    expected_language: str | None = None

    def to_payload(self) -> dict[str, Any]:
        return {
            "case_id": self.case_id,
            "sheet": self.sheet,
            "title": self.title,
            "persona": self.persona,
            "user_turns": self.user_turns,
            "expected_bot_behaviour": self.expected_bot_behaviour,
            "success_criteria": self.success_criteria,
            "red_flags": self.red_flags,
            "expected_language": self.expected_language,
        }


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _nonempty(values: list[Any]) -> list[str]:
    return [_cell_text(value) for value in values if _cell_text(value)]


def _infer_expected_language(*texts: str) -> str | None:
    joined = "\n".join(texts).lower()
    if "websiteaufruf auf deutsch" in joined or "sprache: deutsch" in joined:
        return "de"
    if "websiteaufruf auf englisch" in joined or "sprache: englisch" in joined or "sprache: english" in joined:
        return "en"
    if "englisch" in joined or "english" in joined or "websiteaufruf auf englisch" in joined:
        return "en"
    if "deutsch" in joined or "german" in joined or "websiteaufruf auf deutsch" in joined:
        return "de"
    return None


def _find_header_columns(ws, markers: tuple[str, ...]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            text = _cell_text(cell.value).lower()
            for marker in markers:
                if marker.lower() in text and marker not in columns:
                    columns[marker] = cell.column
    return columns


def _case_blocks(ws) -> list[tuple[int, int]]:
    starts = []
    for row_idx in range(1, ws.max_row + 1):
        value = _cell_text(ws.cell(row_idx, 1).value)
        if value.startswith("TC-"):
            starts.append(row_idx)
    if not starts and ws.max_row:
        starts.append(1)
    blocks = []
    for idx, start in enumerate(starts):
        end = starts[idx + 1] - 1 if idx + 1 < len(starts) else ws.max_row
        blocks.append((start, end))
    return blocks


def parse_uat_excel(path: Path) -> list[UATCase]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse the UAT Excel file.") from exc

    workbook = openpyxl.load_workbook(path, data_only=True)
    cases: list[UATCase] = []

    for ws in workbook.worksheets:
        if ws.title in SKIPPED_SHEETS:
            continue

        header_columns = _find_header_columns(ws, ("Success Criteria", "Red Flags"))
        success_col = header_columns.get("Success Criteria")
        red_col = header_columns.get("Red Flags")

        for start, end in _case_blocks(ws):
            title = _cell_text(ws.cell(start, 1).value)
            case_id_match = re.search(r"(TC-[A-Z0-9-]+)", title)
            case_id = case_id_match.group(1) if case_id_match else f"{ws.title}-{start}"

            flow_row = None
            for row_idx in range(start, end + 1):
                row_values = [_cell_text(ws.cell(row_idx, col).value).lower() for col in range(1, 4)]
                if "dialog flow" in row_values:
                    flow_row = row_idx
                    break

            persona_end = (flow_row - 1) if flow_row else end
            persona = []
            for row_idx in range(start + 1, persona_end + 1):
                text = _cell_text(ws.cell(row_idx, 1).value)
                if text and text.lower() not in {"persona", "dialog flow"} and not text.startswith("TC-"):
                    persona.append(text)

            user_turns = []
            expected = []
            action_texts = []
            for row_idx in range((flow_row or start) + 1, end + 1):
                user_text = _cell_text(ws.cell(row_idx, 2).value)
                bot_text = _cell_text(ws.cell(row_idx, 3).value)
                if user_text:
                    if user_text.startswith("(Aktion)"):
                        action_texts.append(user_text)
                    else:
                        user_turns.append(user_text)
                if bot_text:
                    expected.append(bot_text)

            success = []
            if success_col:
                for row_idx in range(start, end + 1):
                    text = _cell_text(ws.cell(row_idx, success_col).value)
                    if text and not text.lower().startswith(("success criteria", "key features", "positionierung")):
                        success.append(text)

            red_flags = []
            if red_col:
                for row_idx in range(start, end + 1):
                    text = _cell_text(ws.cell(row_idx, red_col).value)
                    if text and not text.lower().startswith("red flags"):
                        red_flags.append(text)

            if not user_turns:
                continue

            expected_language = _infer_expected_language(
                ws.title,
                title,
                "\n".join(persona),
                "\n".join(action_texts),
                "\n".join(user_turns),
                "\n".join(success),
            )
            cases.append(
                UATCase(
                    case_id=case_id,
                    sheet=ws.title,
                    title=title,
                    persona=persona,
                    user_turns=user_turns,
                    expected_bot_behaviour=expected,
                    success_criteria=success,
                    red_flags=red_flags,
                    expected_language=expected_language,
                )
            )

    return cases


def _normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text.lower()).strip()


def _extract_chf_amounts(text: str) -> set[str]:
    amounts = set()
    for match in re.finditer(r"CHF\s*([0-9][0-9'.,\s]*)", text, flags=re.IGNORECASE):
        number = re.sub(r"\D", "", match.group(1))
        if number:
            amounts.add(number)
    return amounts


def _contains_any(text: str, terms: list[str]) -> bool:
    normalized = _normalize_text(text)
    return any(_normalize_text(term) in normalized for term in terms)


def _case_mentions_cost(case: UATCase) -> bool:
    text = "\n".join(case.user_turns + case.expected_bot_behaviour + case.success_criteria)
    return bool(
        re.search(
            r"\b(CHF|cost|costs|fee|fees|tuition|kosten|kostet|geb[uü]hr|studiengeb[uü]hr)\b",
            text,
            flags=re.IGNORECASE,
        )
    )


def _expected_cost_programmes(case: UATCase) -> list[str]:
    """Return the programme costs that should be considered hard truths for this UAT card."""
    sheet = case.sheet.upper()
    if sheet in {"EMBA_DE", "EMBA_DE_EDGE"}:
        return ["emba_hsg"]
    if sheet == "EMBA_DE_CROSS":
        return ["emba_x"]
    if sheet == "IEMBA":
        return ["iemba"]
    if sheet == "IEMBA_EDGE":
        # The card starts as IEMBA but explicitly explores EMBA X as an alternative.
        # Either final cost is acceptable if the answer makes the programme context clear.
        return ["iemba", "emba_x"]
    if sheet == "EMBAX":
        return ["emba_x"]

    rubric = _normalize_text("\n".join(case.expected_bot_behaviour + case.success_criteria + case.user_turns))
    programmes = []
    if "emba x" in rubric or "embax" in rubric:
        programmes.append("emba_x")
    if "iemba" in rubric or "international emba" in rubric:
        programmes.append("iemba")
    if "emba hsg" in rubric or "executive mba hsg" in rubric:
        programmes.append("emba_hsg")
    return programmes


def evaluate_heuristics(case: UATCase, branch_result: dict[str, Any]) -> dict[str, Any]:
    responses = branch_result.get("responses", [])
    joined_response = "\n".join(str(item.get("response", "")) for item in responses)
    joined_rubric = "\n".join(case.expected_bot_behaviour + case.success_criteria)

    checks: list[dict[str, Any]] = []

    if case.expected_language:
        observed_languages = [item.get("language") for item in responses if item.get("language")]
        checks.append(
            {
                "name": "language",
                "expected": case.expected_language,
                "observed": observed_languages,
                "passed": bool(observed_languages) and observed_languages[-1] == case.expected_language,
            }
        )

    programme_expectations = [
        ("emba_x", ["EMBA X", "emba X"], ["emba x", "embax", "eth", "technology", "sustainability"]),
        ("iemba", ["IEMBA", "International EMBA"], ["iemba", "international emba"]),
        ("emba_hsg", ["EMBA HSG", "Executive MBA HSG"], ["emba hsg", "executive mba hsg"]),
    ]
    rubric_lower = joined_rubric.lower()
    for name, labels, triggers in programme_expectations:
        if any(trigger in rubric_lower for trigger in triggers):
            checks.append(
                {
                    "name": f"programme_{name}",
                    "expected": labels,
                    "passed": _contains_any(joined_response, labels),
                }
            )

    if _case_mentions_cost(case):
        expected_programmes = _expected_cost_programmes(case)
        expected_costs = {
            PROGRAMME_COST_TRUTHS[programme]["amount"]
            for programme in expected_programmes
            if programme in PROGRAMME_COST_TRUTHS
        }
        observed_costs = _extract_chf_amounts(joined_response)
        stale_costs = observed_costs & STALE_OR_DISCOUNT_COSTS
        checks.append(
            {
                "name": "hard_cost_amounts",
                "expected": {
                    programme: PROGRAMME_COST_TRUTHS[programme]
                    for programme in expected_programmes
                    if programme in PROGRAMME_COST_TRUTHS
                },
                "observed": sorted(observed_costs),
                "stale_or_discount_amounts_observed": sorted(stale_costs),
                "passed": bool(expected_costs & observed_costs) and not stale_costs,
            }
        )

    booking_terms = ["termin", "handover", "kontaktdaten", "appointment", "slots", "beratung"]
    if any(term in rubric_lower for term in booking_terms):
        booking_seen = any(
            item.get("appointment_requested") or item.get("show_booking_widget")
            for item in responses
        )
        checks.append(
            {
                "name": "booking_or_handover",
                "expected": "booking/contact handover signal",
                "passed": booking_seen or _contains_any(joined_response, ["appointment", "termin", "contact", "kontakt"]),
            }
        )

    red_flag_hits = []
    normalized_response = _normalize_text(joined_response)
    for flag in case.red_flags:
        flag_lower = flag.lower()
        if "bot antwortet auf deutsch" in flag_lower and case.expected_language == "en":
            if any(item.get("language") == "de" for item in responses):
                red_flag_hits.append(flag)
        elif "falsche kostendaten" in flag_lower:
            # Covered by cost_amounts check when the card contains expected costs.
            continue
        elif "spekuliert" in flag_lower:
            speculative_terms = ["garantiert", "definitiv zugelassen", "you will be admitted", "guaranteed admission"]
            if any(term in normalized_response for term in speculative_terms):
                red_flag_hits.append(flag)

    if case.red_flags:
        checks.append(
            {
                "name": "red_flags",
                "expected": "no obvious red flag hit",
                "observed": red_flag_hits,
                "passed": not red_flag_hits,
            }
        )

    passed = sum(1 for check in checks if check.get("passed"))
    score = (passed / len(checks)) if checks else None
    return {
        "score": score,
        "passed_checks": passed,
        "total_checks": len(checks),
        "checks": checks,
    }


def _json_from_model_text(text: str) -> dict[str, Any]:
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, flags=re.DOTALL)
        if not match:
            raise
        return json.loads(match.group(0))


def _mask_stale_cost_expectations(value: Any) -> Any:
    """Remove outdated fee values from UAT expectations before sending them to the judge."""
    if isinstance(value, str):
        stale_pattern = r"(?:CHF\s*)?(72[\s'.,]*500|80[\s'.,]*000|84[\s'.,]*000|99[\s'.,]*000)"
        return re.sub(
            stale_pattern,
            "[STALE_UAT_COST_REMOVED_DO_NOT_USE_AS_EXPECTATION]",
            value,
            flags=re.IGNORECASE,
        )
    if isinstance(value, list):
        return [_mask_stale_cost_expectations(item) for item in value]
    if isinstance(value, dict):
        return {key: _mask_stale_cost_expectations(item) for key, item in value.items()}
    return value


def _openai_client(env_overrides: dict[str, str]):
    try:
        from openai import OpenAI
    except ImportError as exc:
        raise RuntimeError("openai is required for --llm-judge.") from exc

    api_key = (
        env_overrides.get("OPENAI_API_KEY")
        or os.getenv("OPENAI_API_KEY")
        or env_overrides.get("OPENAI_COMPAT_API_KEY")
        or os.getenv("OPENAI_COMPAT_API_KEY")
        or env_overrides.get("OPEN_ROUTER_API_KEY")
        or os.getenv("OPEN_ROUTER_API_KEY")
    )
    base_url = (
        env_overrides.get("OPENAI_COMPAT_BASE_URL")
        or os.getenv("OPENAI_COMPAT_BASE_URL")
        or env_overrides.get("OPEN_ROUTER_BASE_URL")
        or os.getenv("OPEN_ROUTER_BASE_URL")
    )
    if not api_key:
        raise RuntimeError(
            "No judge API key configured. Set OPENAI_API_KEY, OPENAI_COMPAT_API_KEY, or OPEN_ROUTER_API_KEY."
        )
    kwargs = {"api_key": api_key}
    if base_url:
        kwargs["base_url"] = base_url
    return OpenAI(**kwargs)


def _judge_prompt(case: dict[str, Any], branch_id: str, branch_result: dict[str, Any]) -> list[dict[str, str]]:
    responses = branch_result.get("responses", [])
    transcript = []
    for item in responses:
        transcript.append(
            {
                "turn": item.get("turn_index"),
                "user": item.get("query"),
                "assistant": item.get("response"),
                "language": item.get("language"),
                "appointment_requested": item.get("appointment_requested"),
                "show_booking_widget": item.get("show_booking_widget"),
                "error": item.get("error"),
            }
        )

    hard_facts = {
        "programme_costs": {
            truth["label"]: f"CHF {int(truth['amount']):,}".replace(",", "'")
            for truth in PROGRAMME_COST_TRUTHS.values()
        },
        "stale_or_discount_costs_that_should_not_be_present": sorted(STALE_OR_DISCOUNT_COSTS),
    }
    payload = {
        "branch_id": branch_id,
        "scenario": _mask_stale_cost_expectations(case),
        "hard_facts": hard_facts,
        "transcript": transcript,
        "metrics": branch_result.get("metrics", {}),
        "elapsed_s": branch_result.get("elapsed_s"),
        "runner_error": branch_result.get("runner_error", ""),
    }
    system = (
        "You are a strict UAT evaluator for the HSG Executive MBA chatbot. "
        "Evaluate only the observed transcript against the provided scenario, success criteria, red flags, and hard facts. "
        "Hard facts override any conflicting expected behaviour or success criteria from the scenario. "
        "Do not reward verbosity. Penalize incorrect programme fit, stale cost data, missing handover behaviour, language errors, "
        "unsupported claims, and unhelpful or defensive tone. Return valid JSON only."
    )
    user = (
        "Score this branch conversation from 0 to 100. Use the full scenario context. "
        "For fees, treat these as the only correct current values: EMBA HSG CHF 77'500, IEMBA CHF 85'000, emba X CHF 110'000. "
        "If the scenario expects CHF 84'000, CHF 80'000, CHF 72'500, or CHF 99'000, that scenario value is stale and must not be treated as correct. "
        "The branch id is just an identifier; do not prefer any architecture by name. "
        "Return this exact JSON shape:\n"
        "{\n"
        '  "overall_score": number,\n'
        '  "correctness_score": number,\n'
        '  "rag_grounding_score": number,\n'
        '  "conversation_quality_score": number,\n'
        '  "conversion_fit_score": number,\n'
        '  "red_flag_safety_score": number,\n'
        '  "cost_accuracy": "correct|incorrect|not_applicable|unclear",\n'
        '  "passed": boolean,\n'
        '  "verdict": "one concise sentence",\n'
        '  "strengths": ["short bullets"],\n'
        '  "issues": ["short bullets"],\n'
        '  "recommended_followup": "one concise sentence"\n'
        "}\n\n"
        f"UAT payload:\n{json.dumps(payload, ensure_ascii=False, indent=2)}"
    )
    return [{"role": "system", "content": system}, {"role": "user", "content": user}]


def judge_result_with_llm(
    client: Any,
    model: str,
    item: dict[str, Any],
    timeout_s: int,
    **_: Any,
) -> dict[str, Any]:
    messages = _judge_prompt(item["case"], item["branch_id"], item["branch_result"])
    started = time.perf_counter()
    try:
        try:
            completion = client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=0,
                timeout=timeout_s,
                response_format={"type": "json_object"},
            )
        except Exception:
            completion = client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=0,
                timeout=timeout_s,
            )
        content = completion.choices[0].message.content or "{}"
        parsed = _json_from_model_text(content)
        parsed["judge_model"] = model
        parsed["judge_elapsed_s"] = time.perf_counter() - started
        return parsed
    except Exception as exc:
        return {
            "judge_model": model,
            "judge_elapsed_s": time.perf_counter() - started,
            "judge_error": repr(exc),
        }


def add_llm_judgements(
    results: list[dict[str, Any]],
    repo_root: Path,
    model: str,
    timeout_s: int,
) -> list[dict[str, Any]]:
    env_overrides = load_env_file(repo_root / ".env")
    for key in ("LANGSMITH_API_KEY", "LANGSMITH_TRACING", "LANGSMITH_TRACING_V2", "LANGSMITH_PROJECT", "LANGSMITH_ENDPOINT"):
        if env_overrides.get(key):
            os.environ.setdefault(key, env_overrides[key])
    if os.environ.get("LANGSMITH_TRACING") and not os.environ.get("LANGSMITH_TRACING_V2"):
        os.environ["LANGSMITH_TRACING_V2"] = os.environ["LANGSMITH_TRACING"]
    client = _openai_client(env_overrides)
    for index, item in enumerate(results, start=1):
        print(
            f"LLM judge {index}/{len(results)}: {item['branch_id']} {item['case']['case_id']}",
            flush=True,
        )
        metadata = {
            "branch_id": item["branch_id"],
            "branch_ref": item["branch_ref"],
            "case_id": item["case"]["case_id"],
            "sheet": item["case"].get("sheet"),
            "title": item["case"].get("title"),
            "judge_model": model,
            "test_kind": "uat_llm_judge",
        }
        tags = [
            "uat",
            "llm-judge",
            f"branch:{item['branch_id']}",
            f"case:{item['case']['case_id']}",
        ]
        run = None
        if LangSmithRunTree is not None:
            try:
                run = LangSmithRunTree(
                    name="uat_llm_judge",
                    run_type="chain",
                    project_name=os.getenv("LANGSMITH_PROJECT") or "rag2",
                    tags=tags,
                    extra={"metadata": metadata},
                    inputs={
                        **metadata,
                        "model": model,
                        "timeout_s": timeout_s,
                        "branch_elapsed_s": item["branch_result"].get("elapsed_s"),
                        "turn_count": len(item["branch_result"].get("responses", [])),
                        "metrics": item["branch_result"].get("metrics", {}),
                        "transcript": [
                            {
                                "turn_index": response.get("turn_index"),
                                "query": response.get("query"),
                                "response": response.get("response"),
                                "error": response.get("error"),
                            }
                            for response in item["branch_result"].get("responses", [])
                        ],
                    },
                )
                run.post()
            except Exception:
                run = None

        judge_result = judge_result_with_llm(
            client=client,
            model=model,
            item=item,
            timeout_s=timeout_s,
        )
        item["llm_judge"] = judge_result

        if run is not None:
            try:
                run.end(
                    outputs={
                        "overall_score": judge_result.get("overall_score"),
                        "correctness_score": judge_result.get("correctness_score"),
                        "rag_grounding_score": judge_result.get("rag_grounding_score"),
                        "conversation_quality_score": judge_result.get("conversation_quality_score"),
                        "conversion_fit_score": judge_result.get("conversion_fit_score"),
                        "red_flag_safety_score": judge_result.get("red_flag_safety_score"),
                        "cost_accuracy": judge_result.get("cost_accuracy"),
                        "passed": judge_result.get("passed"),
                        "verdict": judge_result.get("verdict"),
                        "strengths": judge_result.get("strengths"),
                        "issues": judge_result.get("issues"),
                        "recommended_followup": judge_result.get("recommended_followup"),
                        "judge_elapsed_s": judge_result.get("judge_elapsed_s"),
                        "judge_error": judge_result.get("judge_error"),
                    },
                    error=judge_result.get("judge_error"),
                )
                run.patch(exclude_inputs=False)
            except Exception:
                pass
    return results


HELPER_CODE = r"""
import json
import os
import sys
import time
import uuid
from datetime import datetime, timezone
from collections import Counter

case = json.loads(sys.stdin.read())
branch_id = case.get("_uat_branch_id", "unknown_branch")
branch_ref = case.get("_uat_branch_ref", "unknown_ref")
event_log_path = case.get("_uat_event_log_path")
metrics = {
    "retrieve_context_calls": 0,
    "retrieve_context_via_tool_calls": 0,
    "tool_calls": Counter(),
    "model_calls": 0,
}
current_turn_index = None

def _preview(value, limit=1200):
    try:
        text = value if isinstance(value, str) else json.dumps(value, ensure_ascii=False, default=str)
    except Exception:
        text = repr(value)
    text = text.replace("\n", "\\n")
    return text[:limit]

def emit_event(event_type, **payload):
    if not event_log_path:
        return
    event = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "event": event_type,
        "branch_id": branch_id,
        "branch_ref": branch_ref,
        "case_id": case.get("case_id"),
        "sheet": case.get("sheet"),
        "title": case.get("title"),
        "turn_index": current_turn_index,
        **payload,
    }
    try:
        os.makedirs(os.path.dirname(event_log_path), exist_ok=True)
        with open(event_log_path, "a", encoding="utf-8") as handle:
            handle.write(json.dumps(event, ensure_ascii=False, default=str) + "\n")
    except Exception:
        pass

emit_event(
    "helper_start",
    user_turn_count=len(case.get("user_turns", [])),
    expected_language=case.get("expected_language"),
)

try:
    from langchain.agents.middleware import wrap_model_call, wrap_tool_call
    from src.rag.middleware import AgentChainMiddleware

    original_tool_wrapper = AgentChainMiddleware._tool_call_wrapper
    original_model_wrapper = AgentChainMiddleware._model_call_wrapper

    def counting_tool_wrapper(request, handler):
        try:
            name = request.tool_call.get("name", "unknown")
            metrics["tool_calls"][name] += 1
        except Exception:
            name = "unknown"
            metrics["tool_calls"]["unknown"] += 1

        tool_call = getattr(request, "tool_call", {}) or {}
        context = getattr(getattr(request, "runtime", None), "context", None)
        agent_name = getattr(context, "agent_name", None)
        call_id = str(uuid.uuid4())
        started = time.perf_counter()
        emit_event(
            "tool_call_start",
            call_id=call_id,
            agent_name=agent_name,
            tool_name=name,
            tool_call_id=tool_call.get("id"),
            tool_args=tool_call.get("args"),
        )
        try:
            response = original_tool_wrapper(request, handler)
            elapsed = time.perf_counter() - started
            emit_event(
                "tool_call_end",
                call_id=call_id,
                agent_name=agent_name,
                tool_name=name,
                tool_call_id=tool_call.get("id"),
                elapsed_s=elapsed,
                output_preview=_preview(getattr(response, "content", response)),
                output_chars=len(str(getattr(response, "content", ""))),
            )
            return response
        except Exception as exc:
            elapsed = time.perf_counter() - started
            emit_event(
                "tool_call_error",
                call_id=call_id,
                agent_name=agent_name,
                tool_name=name,
                tool_call_id=tool_call.get("id"),
                elapsed_s=elapsed,
                error_type=type(exc).__name__,
                error=str(exc),
            )
            raise

    def counting_model_wrapper(request, handler):
        metrics["model_calls"] += 1
        context = getattr(getattr(request, "runtime", None), "context", None)
        agent_name = getattr(context, "agent_name", None)
        model = getattr(request, "model", None)
        model_name = getattr(model, "model_name", None) or getattr(model, "model", None) or repr(model)
        call_id = str(uuid.uuid4())
        started = time.perf_counter()
        emit_event(
            "model_call_start",
            call_id=call_id,
            agent_name=agent_name,
            model_name=model_name,
        )
        try:
            response = original_model_wrapper(request, handler)
            elapsed = time.perf_counter() - started
            result = None
            finish_reason = None
            content_preview = None
            tool_calls = None
            try:
                result = response.result[0]
                finish_reason = result.response_metadata.get("finish_reason")
                content_preview = _preview(getattr(result, "content", ""))
                tool_calls = getattr(result, "tool_calls", None)
            except Exception:
                pass
            emit_event(
                "model_call_end",
                call_id=call_id,
                agent_name=agent_name,
                model_name=model_name,
                elapsed_s=elapsed,
                finish_reason=finish_reason,
                content_preview=content_preview,
                tool_calls=tool_calls,
            )
            return response
        except Exception as exc:
            elapsed = time.perf_counter() - started
            emit_event(
                "model_call_error",
                call_id=call_id,
                agent_name=agent_name,
                model_name=model_name,
                elapsed_s=elapsed,
                error_type=type(exc).__name__,
                error=str(exc),
            )
            raise

    AgentChainMiddleware._tool_wrapper_middleware = wrap_tool_call(counting_tool_wrapper)
    AgentChainMiddleware._model_wrapper_middleware = wrap_model_call(counting_model_wrapper)
except Exception as exc:
    metrics["middleware_patch_error"] = repr(exc)

from src.config import config

try:
    from src.cache.cache import Cache
    Cache.configure(mode="dict", cache=False)
    Cache._instance = None
except Exception:
    pass

try:
    config.chain.EVALUATE_RESPONSE_QUALITY = False
except Exception:
    pass

from src.rag.agent_chain import ExecutiveAgentChain

try:
    from langsmith import tracing_context
    from langsmith.run_trees import RunTree
except Exception:
    tracing_context = None
    RunTree = None

original_retrieve = getattr(ExecutiveAgentChain, "_retrieve_context", None)
if original_retrieve is not None:
    def counted_retrieve(self, *args, **kwargs):
        metrics["retrieve_context_calls"] += 1
        started = time.perf_counter()
        query = kwargs.get("query") if kwargs else None
        program = kwargs.get("program") if kwargs else None
        language_arg = kwargs.get("language") if kwargs else None
        if args:
            query = query if query is not None else (args[0] if len(args) > 0 else None)
            program = program if program is not None else (args[1] if len(args) > 1 else None)
            language_arg = language_arg if language_arg is not None else (args[2] if len(args) > 2 else None)
        call_id = str(uuid.uuid4())
        emit_event(
            "retrieve_context_start",
            call_id=call_id,
            query=_preview(query, 500),
            program=program,
            language=language_arg,
        )
        try:
            result = original_retrieve(self, *args, **kwargs)
            elapsed = time.perf_counter() - started
            emit_event(
                "retrieve_context_end",
                call_id=call_id,
                query=_preview(query, 500),
                program=program,
                language=language_arg,
                elapsed_s=elapsed,
                output_chars=len(result or ""),
                output_preview=_preview(result),
            )
            return result
        except Exception as exc:
            elapsed = time.perf_counter() - started
            emit_event(
                "retrieve_context_error",
                call_id=call_id,
                query=_preview(query, 500),
                program=program,
                language=language_arg,
                elapsed_s=elapsed,
                error_type=type(exc).__name__,
                error=str(exc),
            )
            raise
    ExecutiveAgentChain._retrieve_context = counted_retrieve

original_retrieve_via_tool = getattr(ExecutiveAgentChain, "_retrieve_context_via_tool", None)
if original_retrieve_via_tool is not None:
    def counted_retrieve_via_tool(self, *args, **kwargs):
        metrics["retrieve_context_via_tool_calls"] += 1
        started = time.perf_counter()
        call_id = str(uuid.uuid4())
        emit_event(
            "retrieve_context_via_tool_start",
            call_id=call_id,
            query=_preview(kwargs.get("query") if kwargs else (args[0] if args else None), 500),
            program=kwargs.get("program") if kwargs else (args[1] if len(args) > 1 else None),
            language=kwargs.get("language") if kwargs else (args[2] if len(args) > 2 else None),
        )
        try:
            result = original_retrieve_via_tool(self, *args, **kwargs)
            elapsed = time.perf_counter() - started
            emit_event(
                "retrieve_context_via_tool_end",
                call_id=call_id,
                elapsed_s=elapsed,
                output_chars=len(result or ""),
                output_preview=_preview(result),
            )
            return result
        except Exception as exc:
            elapsed = time.perf_counter() - started
            emit_event(
                "retrieve_context_via_tool_error",
                call_id=call_id,
                elapsed_s=elapsed,
                error_type=type(exc).__name__,
                error=str(exc),
            )
            raise
    ExecutiveAgentChain._retrieve_context_via_tool = counted_retrieve_via_tool

language = case.get("expected_language") or "de"
agent = ExecutiveAgentChain(language=language, session_id=f"uat-{uuid.uuid4()}")

def _run_turn(agent, turn_index, user_turn):
    global current_turn_index
    current_turn_index = turn_index
    turn_start = time.perf_counter()
    emit_event(
        "turn_start",
        query=_preview(user_turn, 1000),
    )
    try:
        response = agent.query(user_turn)
        elapsed = time.perf_counter() - turn_start
        item = {
            "turn_index": turn_index,
            "query": user_turn,
            "elapsed_s": elapsed,
            "response": getattr(response, "response", ""),
            "language": getattr(response, "language", None),
            "appointment_requested": bool(getattr(response, "appointment_requested", False)),
            "show_booking_widget": bool(getattr(response, "show_booking_widget", False)),
            "confidence_fallback": bool(getattr(response, "confidence_fallback", False)),
        }
        emit_event(
            "turn_end",
            elapsed_s=elapsed,
            language=item.get("language"),
            appointment_requested=item.get("appointment_requested"),
            show_booking_widget=item.get("show_booking_widget"),
            confidence_fallback=item.get("confidence_fallback"),
            response_chars=len(item.get("response") or ""),
            response_preview=_preview(item.get("response")),
            metrics_snapshot={**metrics, "tool_calls": dict(metrics["tool_calls"])},
        )
        return item
    except Exception as exc:
        elapsed = time.perf_counter() - turn_start
        item = {
            "turn_index": turn_index,
            "query": user_turn,
            "elapsed_s": elapsed,
            "error": repr(exc),
        }
        emit_event(
            "turn_error",
            elapsed_s=elapsed,
            error_type=type(exc).__name__,
            error=str(exc),
            metrics_snapshot={**metrics, "tool_calls": dict(metrics["tool_calls"])},
        )
        return item
    finally:
        current_turn_index = None

case_metadata = {
    "branch_id": branch_id,
    "branch_ref": branch_ref,
    "case_id": case["case_id"],
    "sheet": case.get("sheet"),
    "title": case.get("title"),
    "expected_language": case.get("expected_language"),
    "test_kind": "uat_branch_comparison",
}
case_tags = [
    "uat",
    "branch-comparison",
    f"branch:{branch_id}",
    f"case:{case['case_id']}",
]

def _run_tree(name, run_type, inputs, parent=None, tags=None, metadata=None):
    if RunTree is None:
        return None
    try:
        if parent is not None:
            run = parent.create_child(
                name=name,
                run_type=run_type,
                inputs=inputs,
                tags=tags,
                extra={"metadata": metadata or {}},
            )
        else:
            run = RunTree(
                name=name,
                run_type=run_type,
                inputs=inputs,
                project_name=os.getenv("LANGSMITH_PROJECT") or "rag2",
                tags=tags,
                extra={"metadata": metadata or {}},
            )
        run.post()
        return run
    except Exception as exc:
        emit_event("langsmith_run_create_error", run_name=name, error_type=type(exc).__name__, error=str(exc))
        return None

def _finish_run_tree(run, outputs=None, error=None):
    if run is None:
        return
    try:
        run.end(outputs=outputs, error=error)
        run.patch(exclude_inputs=False)
    except Exception as exc:
        emit_event(
            "langsmith_run_finish_error",
            run_name=getattr(run, "name", None),
            error_type=type(exc).__name__,
            error=str(exc),
        )

def _run_case(agent):
    responses = []
    case_start = time.perf_counter()
    emit_event("case_start")
    case_run = _run_tree(
        "uat_branch_case",
        "chain",
        inputs={
            "branch_id": branch_id,
            "branch_ref": branch_ref,
            "case_id": case["case_id"],
            "sheet": case.get("sheet"),
            "title": case.get("title"),
            "expected_language": case.get("expected_language"),
            "persona": case.get("persona", []),
            "user_turns": case.get("user_turns", []),
            "success_criteria": case.get("success_criteria", []),
            "red_flags": case.get("red_flags", []),
        },
        tags=case_tags,
        metadata=case_metadata,
    )
    try:
        for turn_index, user_turn in enumerate(case["user_turns"], start=1):
            turn_run = _run_tree(
                "uat_branch_turn",
                "chain",
                inputs={
                    "branch_id": branch_id,
                    "branch_ref": branch_ref,
                    "case_id": case["case_id"],
                    "turn_index": turn_index,
                    "query": user_turn,
                },
                parent=case_run,
                tags=[*case_tags, f"turn:{turn_index}"],
                metadata={**case_metadata, "turn_index": turn_index},
            )
            try:
                if tracing_context is not None and turn_run is not None:
                    with tracing_context(parent=turn_run, metadata={**case_metadata, "turn_index": turn_index}, tags=[*case_tags, f"turn:{turn_index}"]):
                        response_item = _run_turn(agent, turn_index, user_turn)
                else:
                    response_item = _run_turn(agent, turn_index, user_turn)
                _finish_run_tree(
                    turn_run,
                    outputs={
                        "elapsed_s": response_item.get("elapsed_s"),
                        "language": response_item.get("language"),
                        "appointment_requested": response_item.get("appointment_requested"),
                        "show_booking_widget": response_item.get("show_booking_widget"),
                        "confidence_fallback": response_item.get("confidence_fallback"),
                        "error": response_item.get("error"),
                        "response": response_item.get("response"),
                    },
                    error=response_item.get("error"),
                )
            except Exception as exc:
                response_item = {
                    "turn_index": turn_index,
                    "query": user_turn,
                    "elapsed_s": time.perf_counter() - case_start,
                    "error": repr(exc),
                }
                _finish_run_tree(turn_run, outputs=response_item, error=repr(exc))
            responses.append(response_item)
            if response_item.get("error"):
                break
    finally:
        elapsed = time.perf_counter() - case_start
        _finish_run_tree(
            case_run,
            outputs={
                "elapsed_s": elapsed,
                "turn_count": len(responses),
                "metrics": {**metrics, "tool_calls": dict(metrics["tool_calls"])},
                "slowest_turns": sorted(
                    [
                        {
                            "turn_index": item.get("turn_index"),
                            "elapsed_s": item.get("elapsed_s"),
                            "query": str(item.get("query", ""))[:180],
                            "error": item.get("error"),
                        }
                        for item in responses
                    ],
                    key=lambda item: float(item.get("elapsed_s") or 0),
                    reverse=True,
                )[:5],
            },
        )
    elapsed = time.perf_counter() - case_start
    emit_event(
        "case_end",
        elapsed_s=elapsed,
        turn_count=len(responses),
        metrics={**metrics, "tool_calls": dict(metrics["tool_calls"])},
    )
    return responses, elapsed

responses, total_elapsed = _run_case(agent)

print(
    json.dumps(
        {
            "case_id": case["case_id"],
            "responses": responses,
            "elapsed_s": total_elapsed,
            "metrics": {
                **metrics,
                "tool_calls": dict(metrics["tool_calls"]),
            },
        },
        ensure_ascii=False,
    )
)
"""


def run_cmd(args: list[str], cwd: Path, env: dict[str, str] | None = None) -> subprocess.CompletedProcess:
    return subprocess.run(
        args,
        cwd=str(cwd),
        env=env,
        text=True,
        encoding="utf-8",
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )


def load_env_file(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}

    values: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            values[key] = value
    return values


def ensure_worktree(repo_root: Path, worktree_root: Path, branch_ref: str) -> Path:
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", branch_ref.replace("origin/", ""))
    path = worktree_root / safe_name
    if path.exists():
        shutil.rmtree(path)
    result = run_cmd(["git", "worktree", "add", "--detach", str(path), branch_ref], cwd=repo_root)
    if result.returncode != 0:
        raise RuntimeError(f"Failed to create worktree for {branch_ref}:\n{result.stderr}")
    return path


def run_case_in_branch(
    worktree: Path,
    case: UATCase,
    timeout_s: int,
    env_overrides: dict[str, str] | None = None,
    branch_id: str = "unknown_branch",
    branch_ref: str = "unknown_ref",
    event_log_path: Path | None = None,
) -> dict[str, Any]:
    return run_case_in_branch_with_metadata(
        worktree=worktree,
        case=case,
        timeout_s=timeout_s,
        env_overrides=env_overrides,
        branch_id=branch_id,
        branch_ref=branch_ref,
        event_log_path=event_log_path,
    )


def run_case_in_branch_with_metadata(
    worktree: Path,
    case: UATCase,
    timeout_s: int,
    env_overrides: dict[str, str] | None = None,
    branch_id: str = "unknown_branch",
    branch_ref: str = "unknown_ref",
    event_log_path: Path | None = None,
) -> dict[str, Any]:
    env = os.environ.copy()
    if env_overrides:
        env.update(env_overrides)
    if env.get("LANGSMITH_TRACING") and not env.get("LANGSMITH_TRACING_V2"):
        env["LANGSMITH_TRACING_V2"] = env["LANGSMITH_TRACING"]
    env["PYTHONPATH"] = str(worktree)
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"
    payload = {
        **case.to_payload(),
        "_uat_branch_id": branch_id,
        "_uat_branch_ref": branch_ref,
        "_uat_event_log_path": str(event_log_path) if event_log_path else None,
    }
    started = time.perf_counter()
    proc = subprocess.run(
        [sys.executable, "-c", HELPER_CODE],
        input=json.dumps(payload),
        cwd=str(worktree),
        env=env,
        text=True,
        encoding="utf-8",
        errors="replace",
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        timeout=timeout_s,
        check=False,
    )
    elapsed = time.perf_counter() - started
    if proc.returncode != 0:
        return {
            "case_id": case.case_id,
            "elapsed_s": elapsed,
            "responses": [],
            "metrics": {},
            "event_log_path": str(event_log_path) if event_log_path else None,
            "runner_error": proc.stderr.strip() or proc.stdout.strip(),
        }
    try:
        result = json.loads(proc.stdout.strip().splitlines()[-1])
        result["event_log_path"] = str(event_log_path) if event_log_path else None
        return result
    except Exception as exc:
        return {
            "case_id": case.case_id,
            "elapsed_s": elapsed,
            "responses": [],
            "metrics": {},
            "event_log_path": str(event_log_path) if event_log_path else None,
            "runner_error": f"Could not parse helper JSON: {exc}\nSTDOUT:\n{proc.stdout}\nSTDERR:\n{proc.stderr}",
        }


def write_reports(results: list[dict[str, Any]], output_dir: Path) -> tuple[Path, Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    json_path = output_dir / f"uat_branch_comparison_{stamp}.json"
    csv_path = output_dir / f"uat_branch_comparison_{stamp}.csv"
    turns_csv_path = output_dir / f"uat_branch_turns_{stamp}.csv"

    json_path.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")

    fieldnames = [
        "branch_id",
        "branch_ref",
        "case_id",
        "sheet",
        "title",
        "heuristic_score",
        "elapsed_s",
        "turn_count",
        "rag_calls",
        "retrieve_context_via_tool_calls",
        "tool_calls_total",
        "other_tool_calls",
        "model_calls",
        "runner_error",
        "llm_score",
        "llm_passed",
        "llm_cost_accuracy",
        "llm_verdict",
        "llm_issues",
        "last_response_preview",
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for item in results:
            branch_result = item["branch_result"]
            metrics = branch_result.get("metrics", {})
            tool_calls = metrics.get("tool_calls", {}) or {}
            total_tool_calls = sum(tool_calls.values())
            other_tool_calls = {k: v for k, v in tool_calls.items() if k != "retrieve_context"}
            responses = branch_result.get("responses", [])
            last_response = responses[-1].get("response", "") if responses else ""
            writer.writerow(
                {
                    "branch_id": item["branch_id"],
                    "branch_ref": item["branch_ref"],
                    "case_id": item["case"]["case_id"],
                    "sheet": item["case"]["sheet"],
                    "title": item["case"]["title"],
                    "heuristic_score": item["heuristic"].get("score"),
                    "elapsed_s": branch_result.get("elapsed_s"),
                    "turn_count": len(responses),
                    "rag_calls": metrics.get("retrieve_context_calls", 0),
                    "retrieve_context_via_tool_calls": metrics.get("retrieve_context_via_tool_calls", 0),
                    "tool_calls_total": total_tool_calls,
                    "other_tool_calls": json.dumps(other_tool_calls, ensure_ascii=False),
                    "model_calls": metrics.get("model_calls", 0),
                    "runner_error": branch_result.get("runner_error", ""),
                    "llm_score": item.get("llm_judge", {}).get("overall_score"),
                    "llm_passed": item.get("llm_judge", {}).get("passed"),
                    "llm_cost_accuracy": item.get("llm_judge", {}).get("cost_accuracy"),
                    "llm_verdict": item.get("llm_judge", {}).get("verdict", ""),
                    "llm_issues": json.dumps(item.get("llm_judge", {}).get("issues", []), ensure_ascii=False),
                    "last_response_preview": last_response[:240].replace("\n", " "),
                }
            )

    turn_fieldnames = [
        "branch_id",
        "branch_ref",
        "case_id",
        "sheet",
        "title",
        "turn_index",
        "elapsed_s",
        "language",
        "appointment_requested",
        "show_booking_widget",
        "confidence_fallback",
        "query",
        "error",
        "response_preview",
    ]
    with turns_csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=turn_fieldnames)
        writer.writeheader()
        for item in results:
            for response in item["branch_result"].get("responses", []):
                writer.writerow(
                    {
                        "branch_id": item["branch_id"],
                        "branch_ref": item["branch_ref"],
                        "case_id": item["case"]["case_id"],
                        "sheet": item["case"]["sheet"],
                        "title": item["case"]["title"],
                        "turn_index": response.get("turn_index"),
                        "elapsed_s": response.get("elapsed_s"),
                        "language": response.get("language"),
                        "appointment_requested": response.get("appointment_requested"),
                        "show_booking_widget": response.get("show_booking_widget"),
                        "confidence_fallback": response.get("confidence_fallback"),
                        "query": str(response.get("query", "")).replace("\n", " "),
                        "error": response.get("error", ""),
                        "response_preview": str(response.get("response", ""))[:240].replace("\n", " "),
                    }
                )

    return json_path, csv_path, turns_csv_path


def write_llm_summary(results: list[dict[str, Any]], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    path = output_dir / f"uat_llm_judge_summary_{stamp}.md"

    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in results:
        grouped.setdefault(item["branch_id"], []).append(item)

    lines = [
        "# UAT LLM Judge Summary",
        "",
        "## Branch Ranking",
        "",
        "| Branch | Avg LLM Score | Avg Heuristic | Avg Time | RAG Calls | Tool Calls | Judge Errors |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    branch_rows = []
    for branch_id, branch_results in grouped.items():
        llm_scores = [
            float(item.get("llm_judge", {}).get("overall_score"))
            for item in branch_results
            if isinstance(item.get("llm_judge", {}).get("overall_score"), (int, float))
        ]
        heuristic_scores = [
            float(item["heuristic"].get("score"))
            for item in branch_results
            if isinstance(item["heuristic"].get("score"), (int, float))
        ]
        elapsed = [float(item["branch_result"].get("elapsed_s") or 0) for item in branch_results]
        rag_calls = sum(
            int(item["branch_result"].get("metrics", {}).get("retrieve_context_calls", 0))
            for item in branch_results
        )
        tool_calls = sum(
            sum((item["branch_result"].get("metrics", {}).get("tool_calls", {}) or {}).values())
            for item in branch_results
        )
        judge_errors = sum(1 for item in branch_results if item.get("llm_judge", {}).get("judge_error"))
        avg_llm = sum(llm_scores) / len(llm_scores) if llm_scores else 0
        avg_heuristic = sum(heuristic_scores) / len(heuristic_scores) if heuristic_scores else 0
        avg_elapsed = sum(elapsed) / len(elapsed) if elapsed else 0
        branch_rows.append((avg_llm, branch_id, avg_heuristic, avg_elapsed, rag_calls, tool_calls, judge_errors))

    for avg_llm, branch_id, avg_heuristic, avg_elapsed, rag_calls, tool_calls, judge_errors in sorted(branch_rows, reverse=True):
        lines.append(
            f"| {branch_id} | {avg_llm:.1f} | {avg_heuristic:.3f} | {avg_elapsed:.2f}s | "
            f"{rag_calls} | {tool_calls} | {judge_errors} |"
        )

    lines.extend(["", "## Case Winners", ""])
    cases: dict[str, list[dict[str, Any]]] = {}
    for item in results:
        cases.setdefault(item["case"]["case_id"], []).append(item)

    for case_id, case_results in cases.items():
        title = case_results[0]["case"].get("title", case_id)
        ranked = sorted(
            case_results,
            key=lambda item: float(item.get("llm_judge", {}).get("overall_score") or -1),
            reverse=True,
        )
        winner = ranked[0]
        winner_score = winner.get("llm_judge", {}).get("overall_score", "n/a")
        lines.append(f"### {case_id}: {title}")
        lines.append(f"Winner: `{winner['branch_id']}` ({winner_score})")
        for item in ranked:
            judge = item.get("llm_judge", {})
            score = judge.get("overall_score", "n/a")
            verdict = judge.get("verdict") or judge.get("judge_error") or ""
            issues = judge.get("issues") or []
            issue_text = "; ".join(str(issue) for issue in issues[:3])
            suffix = f" Issues: {issue_text}" if issue_text else ""
            lines.append(f"- `{item['branch_id']}`: {score} - {verdict}{suffix}")
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def run_uat_comparison(
    excel_path: Path,
    repo_root: Path,
    output_dir: Path,
    branches: list[tuple[str, str]],
    limit: int | None,
    timeout_s: int,
) -> list[dict[str, Any]]:
    cases = parse_uat_excel(excel_path)
    if limit:
        cases = cases[:limit]
    if not cases:
        raise RuntimeError(f"No UAT cases found in {excel_path}")

    total_runs = len(cases) * len(branches)
    print(
        f"Starting UAT comparison: cases={len(cases)} branches={len(branches)} total_runs={total_runs}",
        flush=True,
    )
    print(f"Excel: {excel_path}", flush=True)
    print(f"Output directory: {output_dir}", flush=True)

    results: list[dict[str, Any]] = []
    output_dir.mkdir(parents=True, exist_ok=True)
    trace_dir = output_dir / "traces"
    trace_dir.mkdir(parents=True, exist_ok=True)
    event_log_path = trace_dir / f"uat_events_{time.strftime('%Y%m%d_%H%M%S')}.jsonl"
    print(f"Event trace log: {event_log_path}", flush=True)
    with tempfile.TemporaryDirectory(prefix="hsg-rag-uat-worktrees-") as tmp:
        worktree_root = Path(tmp)
        env_overrides = load_env_file(repo_root / ".env")
        print(f"Creating temporary worktrees under: {worktree_root}", flush=True)
        worktrees = {}
        for branch_id, branch_ref in branches:
            print(f"Creating worktree for {branch_id}: {branch_ref}", flush=True)
            worktrees[branch_id] = ensure_worktree(repo_root, worktree_root, branch_ref)
            print(f"Ready worktree for {branch_id}: {worktrees[branch_id]}", flush=True)

        completed_runs = 0
        for case_index, case in enumerate(cases, start=1):
            print(
                f"Case {case_index}/{len(cases)} {case.case_id}: {case.title}",
                flush=True,
            )
            for branch_id, branch_ref in branches:
                completed_runs += 1
                started = time.perf_counter()
                print(
                    f"[{completed_runs}/{total_runs}] START {branch_id} {case.case_id}",
                    flush=True,
                )
                branch_result = run_case_in_branch(
                    worktrees[branch_id],
                    case,
                    timeout_s=timeout_s,
                    env_overrides=env_overrides,
                    branch_id=branch_id,
                    branch_ref=branch_ref,
                    event_log_path=event_log_path,
                )
                heuristic = evaluate_heuristics(case, branch_result)
                elapsed = time.perf_counter() - started
                status = "ERROR" if branch_result.get("runner_error") else "OK"
                score = heuristic.get("score")
                rag_calls = branch_result.get("metrics", {}).get("retrieve_context_calls", 0)
                model_calls = branch_result.get("metrics", {}).get("model_calls", 0)
                print(
                    f"[{completed_runs}/{total_runs}] END {branch_id} {case.case_id} "
                    f"status={status} elapsed_s={elapsed:.1f} score={score} "
                    f"rag_calls={rag_calls} model_calls={model_calls}",
                    flush=True,
                )
                results.append(
                    {
                        "branch_id": branch_id,
                        "branch_ref": branch_ref,
                        "case": case.to_payload(),
                        "branch_result": branch_result,
                        "heuristic": heuristic,
                    }
                )

    write_reports(results, output_dir)
    print(f"Local event trace log: {event_log_path}")
    return results


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare HSG RAG UAT cases across branches.")
    parser.add_argument("--excel", type=Path, default=Path(os.getenv("UAT_EXCEL", DEFAULT_EXCEL)))
    parser.add_argument("--repo-root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--output-dir", type=Path, default=Path(os.getenv("UAT_OUTPUT_DIR", "/tmp/hsg_rag_uat_results")))
    parser.add_argument("--limit", type=int, default=int(os.getenv("UAT_LIMIT", "0")) or None)
    parser.add_argument("--timeout-s", type=int, default=int(os.getenv("UAT_TIMEOUT_S", "180")))
    parser.add_argument("--llm-judge", action="store_true", default=os.getenv("UAT_LLM_JUDGE") == "1")
    parser.add_argument("--judge-model", default=DEFAULT_JUDGE_MODEL)
    parser.add_argument("--judge-timeout-s", type=int, default=int(os.getenv("UAT_JUDGE_TIMEOUT_S", "120")))
    parser.add_argument(
        "--judge-existing-report",
        type=Path,
        default=None,
        help="Evaluate an existing uat_branch_comparison_*.json report without rerunning branch conversations.",
    )
    parser.add_argument(
        "--branch",
        action="append",
        default=[],
        help="Optional branch mapping as id=ref. Can be supplied multiple times.",
    )
    return parser.parse_args(argv)


def branch_args_to_pairs(values: list[str]) -> list[tuple[str, str]]:
    if not values:
        return DEFAULT_BRANCHES
    pairs = []
    for value in values:
        if "=" not in value:
            raise ValueError(f"Branch must be supplied as id=ref, got: {value}")
        branch_id, branch_ref = value.split("=", 1)
        pairs.append((branch_id.strip(), branch_ref.strip()))
    return pairs


def print_summary(results: list[dict[str, Any]], output_dir: Path) -> None:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in results:
        grouped.setdefault(item["branch_id"], []).append(item)

    print("\nUAT branch comparison summary")
    print(f"Output directory: {output_dir}")
    for branch_id, branch_results in grouped.items():
        scores = [
            item["heuristic"].get("score")
            for item in branch_results
            if item["heuristic"].get("score") is not None
        ]
        elapsed = [float(item["branch_result"].get("elapsed_s") or 0) for item in branch_results]
        rag_calls = [
            int(item["branch_result"].get("metrics", {}).get("retrieve_context_calls", 0))
            for item in branch_results
        ]
        errors = [item for item in branch_results if item["branch_result"].get("runner_error")]
        avg_score = sum(scores) / len(scores) if scores else None
        avg_elapsed = sum(elapsed) / len(elapsed) if elapsed else 0
        print(
            f"- {branch_id}: cases={len(branch_results)} "
            f"avg_score={avg_score if avg_score is not None else 'n/a'} "
            f"avg_elapsed_s={avg_elapsed:.2f} "
            f"rag_calls={sum(rag_calls)} "
            f"errors={len(errors)}"
        )
        llm_scores = [
            float(item.get("llm_judge", {}).get("overall_score"))
            for item in branch_results
            if isinstance(item.get("llm_judge", {}).get("overall_score"), (int, float))
        ]
        if llm_scores:
            print(f"  llm_avg_score={sum(llm_scores) / len(llm_scores):.1f}")

    slow_turns = []
    for item in results:
        for response in item["branch_result"].get("responses", []):
            slow_turns.append(
                {
                    "branch_id": item["branch_id"],
                    "case_id": item["case"]["case_id"],
                    "turn_index": response.get("turn_index"),
                    "elapsed_s": float(response.get("elapsed_s") or 0),
                    "query": str(response.get("query", ""))[:90].replace("\n", " "),
                }
            )
    if slow_turns:
        print("\nSlowest UAT turns")
        for turn in sorted(slow_turns, key=lambda item: item["elapsed_s"], reverse=True)[:10]:
            print(
                f"- {turn['branch_id']} {turn['case_id']} turn={turn['turn_index']} "
                f"elapsed_s={turn['elapsed_s']:.2f} query={turn['query']}"
            )

    event_log_paths = sorted(
        {
            item["branch_result"].get("event_log_path")
            for item in results
            if item["branch_result"].get("event_log_path")
        }
    )
    slow_events = []
    for raw_path in event_log_paths:
        path = Path(raw_path)
        if not path.exists():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            try:
                event = json.loads(line)
            except json.JSONDecodeError:
                continue
            if event.get("elapsed_s") is None:
                continue
            if event.get("event") not in {
                "turn_end",
                "model_call_end",
                "tool_call_end",
                "retrieve_context_end",
                "retrieve_context_via_tool_end",
                "case_end",
            }:
                continue
            slow_events.append(event)

    if slow_events:
        print("\nSlowest internal UAT events")
        for event in sorted(slow_events, key=lambda item: float(item.get("elapsed_s") or 0), reverse=True)[:15]:
            subject = (
                event.get("model_name")
                or event.get("tool_name")
                or event.get("program")
                or event.get("event")
            )
            print(
                f"- {event.get('branch_id')} {event.get('case_id')} turn={event.get('turn_index')} "
                f"{event.get('event')} elapsed_s={float(event.get('elapsed_s') or 0):.2f} "
                f"agent={event.get('agent_name')} subject={subject} "
                f"query={str(event.get('query') or '')[:90]}"
            )
        if event_log_paths:
            print(f"\nLocal event trace log: {event_log_paths[-1]}")


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    branches = branch_args_to_pairs(args.branch)
    if args.judge_existing_report:
        results = json.loads(args.judge_existing_report.read_text(encoding="utf-8"))
    else:
        results = run_uat_comparison(
            excel_path=args.excel,
            repo_root=args.repo_root,
            output_dir=args.output_dir,
            branches=branches,
            limit=args.limit,
            timeout_s=args.timeout_s,
        )

    if args.llm_judge:
        results = add_llm_judgements(
            results=results,
            repo_root=args.repo_root,
            model=args.judge_model,
            timeout_s=args.judge_timeout_s,
        )
        json_path, csv_path, turns_csv_path = write_reports(results, args.output_dir)
        summary_path = write_llm_summary(results, args.output_dir)
        print(f"\nLLM judge JSON report: {json_path}")
        print(f"LLM judge CSV report: {csv_path}")
        print(f"Turn timing CSV report: {turns_csv_path}")
        print(f"LLM judge summary: {summary_path}")

    print_summary(results, args.output_dir)
    return 0


@pytest.mark.integration
@pytest.mark.network
def test_uat_branch_comparison_opt_in():
    if os.getenv("RUN_UAT_BRANCH_COMPARISON") != "1":
        pytest.skip("Set RUN_UAT_BRANCH_COMPARISON=1 to run cross-branch UAT comparison.")
    args = parse_args([])
    results = run_uat_comparison(
        excel_path=args.excel,
        repo_root=args.repo_root,
        output_dir=args.output_dir,
        branches=DEFAULT_BRANCHES,
        limit=args.limit,
        timeout_s=args.timeout_s,
    )
    assert results


if __name__ == "__main__":
    raise SystemExit(main())
