"""
UAT acceptance gate: run the chatbot through workbook scenarios and judge the
full conversation with an LLM.

Opt-in because this calls live LLMs and Weaviate:
    RUN_UAT_LLM_JUDGE=1 pytest tests/test_uat_llm_judge.py -v -s
"""
from __future__ import annotations

import json
import os
import re
import sys
import time
import uuid
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.config import config
from src.rag.agent_chain import ExecutiveAgentChain


OPENROUTER_BASE_URL = "https://openrouter.ai/api/v1"


def _use_openrouter() -> bool:
    """Route the LLM judge through OpenRouter by default (opt out with
    UAT_USE_OPENROUTER=0).

    The chatbot's own models already run through OpenRouter via config.py; this
    keeps the judge consistent so the UAT does not depend on a funded OpenAI
    account.
    """
    if os.getenv("UAT_USE_OPENROUTER") == "0":
        return False
    return bool(os.getenv("OPEN_ROUTER_API_KEY"))


DEFAULT_EXCEL = Path(__file__).resolve().parent / "fixtures" / "UAT.xlsx"
SKIPPED_SHEETS = {"About", "TestProtocol", "Reporting"}
DEFAULT_JUDGE_MODEL = os.getenv("UAT_JUDGE_MODEL") or (
    "openai/gpt-4o-mini" if _use_openrouter() else "gpt-4o-mini"
)
AVERAGE_MIN_ACCEPTABLE_SCORE = float(
    os.getenv("UAT_AVERAGE_MIN_SCORE", os.getenv("UAT_MIN_SCORE", "8.5"))
)
CASE_MIN_ACCEPTABLE_SCORE = float(os.getenv("UAT_CASE_MIN_SCORE", "8.0"))


@dataclass(frozen=True)
class UATCase:
    case_id: str
    sheet: str
    title: str
    persona: list[str] = field(default_factory=list)
    user_turns: list[str] = field(default_factory=list)
    expected_bot_behaviour: list[str] = field(default_factory=list)
    success_criteria: list[str] = field(default_factory=list)
    red_flags: list[str] = field(default_factory=list)
    expected_language: str | None = None
    greeting_expected: bool = False
    mixed_language_clarification_expected: bool = False

    def to_payload(self) -> dict[str, Any]:
        return {
            "case_id": self.case_id,
            "sheet": self.sheet,
            "title": self.title,
            "persona": self.persona,
            "user_turns": self.user_turns,
            "expected_bot_behaviour": self.expected_bot_behaviour,
            "success_criteria": self.success_criteria,
            "red_flags": self.red_flags,
            "expected_language": self.expected_language,
            "greeting_expected": self.greeting_expected,
            "mixed_language_clarification_expected": self.mixed_language_clarification_expected,
        }


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _infer_expected_language(*texts: str) -> str | None:
    joined = "\n".join(texts).lower()
    if "websiteaufruf auf deutsch" in joined or "sprache: deutsch" in joined:
        return "de"
    if (
        "websiteaufruf auf englisch" in joined
        or "sprache: englisch" in joined
        or "sprache: english" in joined
    ):
        return "en"
    if "deutsch" in joined or "german" in joined:
        return "de"
    if "englisch" in joined or "english" in joined:
        return "en"
    return None


def _is_greeting_expectation(text: str) -> bool:
    normalized = text.strip().lower()
    normalized = re.sub(r"^[\s→\-–—>*•]+", "", normalized).strip()
    return normalized.startswith(("begrüßung", "begruessung", "greeting"))


def _expects_mixed_language_clarification(
    title: str,
    user_turns: list[str],
    expected_bot_behaviour: list[str],
) -> bool:
    joined = "\n".join([title, *user_turns, *expected_bot_behaviour]).lower()
    return (
        "multiple languages" in joined
        or "mehrere sprachen" in joined
        or "dominante sprache" in joined
        or "english or german" in joined
        or "deutsch oder englisch" in joined
    )


def _case_blocks(ws) -> list[tuple[int, int]]:
    starts = []
    for row_idx in range(1, ws.max_row + 1):
        if _cell_text(ws.cell(row_idx, 1).value).startswith("TC-"):
            starts.append(row_idx)

    blocks = []
    for idx, start in enumerate(starts):
        end = starts[idx + 1] - 1 if idx + 1 < len(starts) else ws.max_row
        blocks.append((start, end))
    return blocks


def _find_header_columns(ws, markers: tuple[str, ...]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            text = _cell_text(cell.value).lower()
            for marker in markers:
                if marker.lower() in text and marker not in columns:
                    columns[marker] = cell.column
    return columns


def parse_uat_excel(path: Path = DEFAULT_EXCEL) -> list[UATCase]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse tests/fixtures/UAT.xlsx") from exc

    workbook = openpyxl.load_workbook(path, data_only=True)
    cases: list[UATCase] = []

    for ws in workbook.worksheets:
        if ws.title in SKIPPED_SHEETS:
            continue

        header_columns = _find_header_columns(ws, ("Success Criteria", "Red Flags"))
        success_col = header_columns.get("Success Criteria")
        red_col = header_columns.get("Red Flags")

        for start, end in _case_blocks(ws):
            title = _cell_text(ws.cell(start, 1).value)
            case_id_match = re.search(r"(TC-[A-Z0-9-]+)", title)
            case_id = case_id_match.group(1) if case_id_match else f"{ws.title}-{start}"

            flow_row = None
            for row_idx in range(start, end + 1):
                row_values = [_cell_text(ws.cell(row_idx, col).value).lower() for col in range(1, 4)]
                if "dialog flow" in row_values:
                    flow_row = row_idx
                    break

            persona = []
            persona_end = (flow_row - 1) if flow_row else end
            for row_idx in range(start + 1, persona_end + 1):
                text = _cell_text(ws.cell(row_idx, 1).value)
                if text and text.lower() not in {"persona", "dialog flow"} and not text.startswith("TC-"):
                    persona.append(text)

            user_turns = []
            expected = []
            greeting_expected = False
            action_texts = []
            for row_idx in range((flow_row or start) + 1, end + 1):
                user_text = _cell_text(ws.cell(row_idx, 2).value)
                bot_text = _cell_text(ws.cell(row_idx, 3).value)
                if user_text:
                    if user_text.startswith("(Aktion)"):
                        action_texts.append(user_text)
                    else:
                        user_turns.append(user_text)
                if bot_text:
                    if _is_greeting_expectation(bot_text):
                        greeting_expected = True
                    else:
                        expected.append(bot_text)

            if not user_turns:
                continue

            success = []
            if success_col:
                for row_idx in range(start, end + 1):
                    text = _cell_text(ws.cell(row_idx, success_col).value)
                    if text and not text.lower().startswith(("success criteria", "key features", "positionierung")):
                        success.append(text)

            red_flags = []
            if red_col:
                for row_idx in range(start, end + 1):
                    text = _cell_text(ws.cell(row_idx, red_col).value)
                    if text and not text.lower().startswith("red flags"):
                        red_flags.append(text)

            expected_language = _infer_expected_language(
                ws.title,
                title,
                "\n".join(persona),
                "\n".join(action_texts),
                "\n".join(user_turns),
                "\n".join(success),
            )
            cases.append(
                UATCase(
                    case_id=case_id,
                    sheet=ws.title,
                    title=title,
                    persona=persona,
                    user_turns=user_turns,
                    expected_bot_behaviour=expected,
                    success_criteria=success,
                    red_flags=red_flags,
                    expected_language=expected_language,
                    greeting_expected=greeting_expected,
                    mixed_language_clarification_expected=_expects_mixed_language_clarification(
                        title,
                        user_turns,
                        expected,
                    ),
                )
            )

    return cases


def _selected_cases() -> list[UATCase]:
    cases = parse_uat_excel()
    limit = int(os.getenv("UAT_LIMIT", "0") or "0")
    if limit > 0:
        return cases[:limit]
    return cases


def _parse_reference_date(raw: str | None) -> date:
    if not raw:
        return date.today()
    return date.fromisoformat(raw[:10])


def _deadline_status(deadline: str | None, reference_date: date) -> str:
    if not deadline:
        return "unknown"
    deadline_date = date.fromisoformat(deadline[:10])
    if deadline_date < reference_date:
        return "passed"
    if deadline_date == reference_date:
        return "due_today"
    return "future"


def _tuition_with_deadline_status(tuition: dict[str, Any], reference_date: date) -> dict[str, Any]:
    enriched = {
        "first_deadline": dict(tuition["first_deadline"]),
        "final_deadline": dict(tuition["final_deadline"]),
        "note": tuition.get("note", {}),
    }

    first = enriched["first_deadline"]
    final = enriched["final_deadline"]
    first["deadline_status"] = _deadline_status(first.get("deadline"), reference_date)
    final["deadline_status"] = _deadline_status(final.get("deadline"), reference_date)

    applicable = first if first["deadline_status"] in {"future", "due_today"} else final
    applicable_key = (
        "first_deadline"
        if first["deadline_status"] in {"future", "due_today"}
        else "final_deadline"
    )
    enriched["applicable_today"] = {
        "deadline_type": applicable_key,
        "deadline": applicable.get("deadline"),
        "deadline_status": applicable.get("deadline_status"),
        "fee": applicable.get("fee"),
    }
    return enriched


def _hard_facts() -> dict[str, Any]:
    path = Path(config.paths.DATA) / "database" / "programme_facts.json"
    with path.open(encoding="utf-8") as f:
        facts = json.load(f)

    reference_date = _parse_reference_date(facts.get("generated_at"))
    programmes = {}
    for key, programme in facts["programmes"].items():
        tuition = programme["tuition_chf"]
        programmes[key] = {
            "official_name": programme["official_name"],
            "language": programme["language"],
            "programme_start": programme["programme_start"],
            "duration": programme["duration"],
            "ects_credits": programme["ects_credits"],
            "structure": programme["structure"],
            "locations": programme["locations"],
            "tuition_chf": _tuition_with_deadline_status(tuition, reference_date),
            "advisor": programme["advisor"],
        }
    return {
        "generated_at": facts.get("generated_at"),
        "reference_date": reference_date.isoformat(),
        "programmes": programmes,
    }


def _run_chatbot_case(case: UATCase) -> list[dict[str, Any]]:
    app_language = case.expected_language or "en"
    chain = ExecutiveAgentChain(
        language=app_language,
        session_id=f"uat-{case.case_id}-{uuid.uuid4()}",
    )

    greeting_started = time.perf_counter()
    transcript = [
        {
            "turn": 0,
            "message_type": "greeting",
            "user": None,
            "assistant": chain.generate_greeting(),
            "additional_details": "",
            "language": app_language,
            "appointment_requested": False,
            "relevant_programs": [],
            "elapsed_s": round(time.perf_counter() - greeting_started, 3),
        }
    ]
    for turn_index, query in enumerate(case.user_turns, start=1):
        started = time.perf_counter()
        result = chain.query(query)
        transcript.append(
            {
                "turn": turn_index,
                "user": query,
                "assistant": result.response,
                "additional_details": result.additional_details or "",
                "language": result.language,
                "appointment_requested": result.appointment_requested,
                "relevant_programs": result.relevant_programs,
                "elapsed_s": round(time.perf_counter() - started, 3),
            }
        )
    return transcript


def _openai_client():
    try:
        from openai import OpenAI
    except ImportError as exc:
        raise RuntimeError("openai is required for the UAT LLM judge.") from exc

    if _use_openrouter():
        return OpenAI(
            api_key=os.environ["OPEN_ROUTER_API_KEY"],
            base_url=OPENROUTER_BASE_URL,
            timeout=float(os.getenv("UAT_JUDGE_TIMEOUT_S", "120")),
        )

    api_key = os.getenv("OPENAI_API_KEY") or os.getenv("OPENAI_COMPAT_API_KEY")
    base_url = os.getenv("OPENAI_COMPAT_BASE_URL")
    if not api_key:
        raise RuntimeError("Set OPENAI_API_KEY for RUN_UAT_LLM_JUDGE=1.")

    kwargs: dict[str, Any] = {
        "api_key": api_key,
        "timeout": float(os.getenv("UAT_JUDGE_TIMEOUT_S", "120")),
    }
    if base_url:
        kwargs["base_url"] = base_url
    return OpenAI(**kwargs)


def _judge_case(case: UATCase, transcript: list[dict[str, Any]]) -> dict[str, Any]:
    payload = {
        "scenario": case.to_payload(),
        "hard_facts_current_source_of_truth": _hard_facts(),
        "transcript": transcript,
        "minimum_acceptable_average_score": AVERAGE_MIN_ACCEPTABLE_SCORE,
        "minimum_acceptable_case_score": CASE_MIN_ACCEPTABLE_SCORE,
    }
    messages = [
        {
            "role": "system",
            "content": (
                "You are a strict but fair UAT judge for the HSG Executive Education chatbot. "
                "Evaluate the whole transcript against the scenario, success criteria, red flags, "
                "and current hard facts. Current hard facts override stale scenario wording. "
                "Use hard_facts_current_source_of_truth.reference_date as the reference date for "
                "deadline-sensitive tuition. The tuition_chf.applicable_today object identifies "
                "which fee and deadline apply on the reference date; same-day deadlines are still "
                "available and are marked due_today. Prefer applicable_today over stale scenario "
                "wording or your own date arithmetic. "
                "Do not grade any legacy booking-widget visibility flag; that flag is not part "
                "of the current acceptance criteria. For appointment/booking/Termin scenarios, "
                "judge only the user-facing behaviour: whether the assistant acknowledges the "
                "user's intent and explains how the user can make a Termin, such as using the "
                "booking section at the bottom of the page or following the correct advisor/contact "
                "path. Do not fail a case because a hidden widget flag is absent or false. "
                "Do not require exact phrasing. Reward helpful, grounded, concise advisory answers. "
                "For scenarios marked mixed_language_clarification_expected, expect and accept a "
                "concise English clarification that asks whether to continue in English or German, "
                "even when the app greeting language is German. Every transcript starts with a "
                "bot-only greeting (turn 0) generated for the app language inferred from the "
                "workbook. If scenario.greeting_expected is true, turn 0 satisfies that workbook "
                "greeting expectation when it uses scenario.expected_language; do not require exact "
                "greeting wording and do not require later assistant replies to greet again. "
                "Penalize wrong programme facts, wrong language, unsupported promises, hidden/internal "
                "architecture talk, missed booking or handover behavior, and rude tone. "
                "Return only valid JSON."
            ),
        },
        {
            "role": "user",
            "content": (
                "Score this UAT transcript from 0 to 10. "
                "Return JSON with keys: overall_score, passed, verdict, strengths, issues, "
                "criteria_met, criteria_missed, language_ok, factuality_ok, tone_ok.\n\n"
                f"{json.dumps(payload, ensure_ascii=False, indent=2)}"
            ),
        },
    ]

    client = _openai_client()
    completion = client.chat.completions.create(
        model=DEFAULT_JUDGE_MODEL,
        messages=messages,
        temperature=0,
        response_format={"type": "json_object"},
    )
    content = completion.choices[0].message.content or "{}"
    return json.loads(content)


def _result_dir() -> Path:
    return Path(os.getenv("UAT_RESULTS_DIR", "uat-results"))


def _write_case_result(
    case: UATCase,
    transcript: list[dict[str, Any]],
    judgement: dict[str, Any],
    score: float,
    passed: bool,
    error: str | None = None,
) -> None:
    result_dir = _result_dir()
    result_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "case_id": case.case_id,
        "title": case.title,
        "score": score,
        "minimum_score": CASE_MIN_ACCEPTABLE_SCORE,
        "average_min_score": AVERAGE_MIN_ACCEPTABLE_SCORE,
        "case_min_score": CASE_MIN_ACCEPTABLE_SCORE,
        "passed": passed,
        "verdict": judgement.get("verdict"),
        "issues": judgement.get("issues", []),
        "criteria_missed": judgement.get("criteria_missed", []),
        "judgement": judgement,
        "transcript": transcript,
        "error": error,
    }
    path = result_dir / f"{case.case_id}.json"
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


UAT_CASES = _selected_cases()


def _run_and_judge_case(case: UATCase) -> dict[str, Any]:
    transcript = []
    judgement: dict[str, Any] = {}
    try:
        transcript = _run_chatbot_case(case)
        judgement = _judge_case(case, transcript)
        score = float(judgement.get("overall_score", 0))
    except Exception as exc:
        score = 0.0
        judgement = {
            "passed": False,
            "verdict": f"UAT runner or judge error: {exc}",
            "issues": [repr(exc)],
            "criteria_missed": ["UAT case could not complete"],
        }
        _write_case_result(case, transcript, judgement, score, passed=False, error=repr(exc))
        return {
            "case": case,
            "score": score,
            "passed": False,
            "judgement": judgement,
            "transcript": transcript,
            "error": repr(exc),
        }

    passed = score >= CASE_MIN_ACCEPTABLE_SCORE
    _write_case_result(case, transcript, judgement, score, passed=passed)
    return {
        "case": case,
        "score": score,
        "passed": passed,
        "judgement": judgement,
        "transcript": transcript,
        "error": None,
    }


@pytest.mark.network
@pytest.mark.integration
@pytest.mark.skipif(
    os.getenv("RUN_UAT_LLM_JUDGE") != "1",
    reason="Set RUN_UAT_LLM_JUDGE=1 to run live UAT conversations and LLM judge.",
)
def test_uat_average_score_passes_llm_judge():
    assert UAT_CASES, "Expected at least one UAT case."

    results = [_run_and_judge_case(case) for case in UAT_CASES]
    scores = [float(result["score"]) for result in results]
    average_score = sum(scores) / len(scores)
    errored = [result for result in results if result.get("error")]
    below_case_floor = [
        result for result in results if float(result["score"]) < CASE_MIN_ACCEPTABLE_SCORE
    ]

    failure_summary = "\n".join(
        (
            f"- {result['case'].case_id}: score {result['score']:.1f}/{CASE_MIN_ACCEPTABLE_SCORE}; "
            f"{result['judgement'].get('verdict')}"
        )
        for result in below_case_floor
    )

    error_summary = "\n".join(
        f"- {result['case'].case_id}: {result.get('error')}" for result in errored
    )

    assert not errored, f"\nUAT runner or judge errors:\n{error_summary}"
    assert not below_case_floor, (
        f"\nUAT cases below minimum case score {CASE_MIN_ACCEPTABLE_SCORE}:"
        f"\n{failure_summary or 'None'}"
    )
    assert average_score >= AVERAGE_MIN_ACCEPTABLE_SCORE, (
        f"\nUAT average score {average_score:.2f} is below minimum "
        f"{AVERAGE_MIN_ACCEPTABLE_SCORE}."
    )


def test_uat_workbook_contains_cases():
    assert len(parse_uat_excel()) > 0


def test_uat_workbook_languages_match_app_selection():
    languages = {case.case_id: case.expected_language for case in parse_uat_excel()}

    assert languages == {
        "TC-EMBA-01": "de",
        "TC-EMBA-02": "de",
        "TC-EMBA-03": "de",
        "TC-IEMBA-01": "en",
        "TC-IEMBA-02": "en",
        "TC-EMBAX-01": "de",
        "TC-EDGE-01": "de",
        "TC-EDGE-02": "de",
        "TC-EDGE-03": "de",
        "TC-EDGE-04": "de",
        "TC-EDGE-05": "de",
    }


def test_uat_expected_behaviour_excludes_greeting_rows():
    cases = parse_uat_excel()

    assert any(case.greeting_expected for case in cases)
    assert all(
        "begrüßung" not in "\n".join(case.expected_bot_behaviour).lower()
        for case in cases
    )


def test_uat_mixed_language_case_expects_english_clarification():
    cases = {case.case_id: case for case in parse_uat_excel()}

    assert cases["TC-EDGE-05"].expected_language == "de"
    assert cases["TC-EDGE-05"].mixed_language_clarification_expected is True


def test_hard_facts_include_deadline_status_and_applicable_fee():
    facts = _hard_facts()

    assert facts["reference_date"] == "2026-06-30"
    assert facts["programmes"]["iemba"]["tuition_chf"]["first_deadline"]["deadline_status"] == "passed"
    assert facts["programmes"]["iemba"]["tuition_chf"]["final_deadline"]["deadline_status"] == "due_today"
    assert facts["programmes"]["iemba"]["tuition_chf"]["applicable_today"]["fee"] == 85000
    assert facts["programmes"]["emba_x"]["tuition_chf"]["applicable_today"]["fee"] == 99000


def test_uat_transcript_starts_with_bot_greeting(monkeypatch):
    class FakeResult:
        response = "Programme answer"
        additional_details = None
        language = "de"
        appointment_requested = False
        relevant_programs = []

    class FakeChain:
        def __init__(self, language: str, session_id: str):
            self.language = language

        def generate_greeting(self) -> str:
            return f"Greeting in {self.language}"

        def query(self, query: str) -> FakeResult:
            return FakeResult()

    monkeypatch.setattr(sys.modules[__name__], "ExecutiveAgentChain", FakeChain)
    case = UATCase(
        case_id="TC-GREETING",
        sheet="Example",
        title="Greeting test",
        user_turns=["Hallo"],
        expected_language="de",
    )

    transcript = _run_chatbot_case(case)

    assert transcript[0]["turn"] == 0
    assert transcript[0]["message_type"] == "greeting"
    assert transcript[0]["user"] is None
    assert transcript[0]["assistant"] == "Greeting in de"
    assert transcript[0]["language"] == "de"
    assert transcript[1]["turn"] == 1
    assert transcript[1]["user"] == "Hallo"
